# VIEWS - DE CORE
from django.db import models
from django.shortcuts import render,  redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm
from django.urls import reverse
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.conf import settings
# Create your views here.
from core.models import PerfilUsuario, EmpresaDB, Empresa
from timbres.models import MovimientoTimbresGlobal
from core._thread_locals import get_current_tenant, get_current_empresa_id, set_current_tenant
from core.db_router import set_current_tenant_connection
from django.db import connections
from django.core.exceptions import PermissionDenied
from functools import wraps
from core._thread_locals import _thread_locals
from django.utils.timezone import now, localtime
from decouple import config
from core.mixins import TenantRequiredMixin

def require_tenant_connection(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        # Verificar que el usuario esté autenticado
        if not request.user.is_authenticated:
            return redirect('core:login')

        # Validamos que la sesión tenga la clave 'tenant'
        if not get_current_tenant() or not get_current_empresa_id():
            messages.error(request, "La sesión del tenant es inválida o ha expirado.")
            logout(request)
            return redirect('core:login')

        # Verificamos si 'tenant' está en la sesión
        if request.session.get('alias_tenant') != 'tenant':
            messages.error(request, "La sesión del tenant es inválida o ha expirado.")
            logout(request)
            return redirect('core:login')

        # Continuar con la vista original si todo es válido
        return view_func(request, *args, **kwargs)

    return wrapper


@login_required
@require_tenant_connection
def inicio(request):
    try:
        # Leemos el perfil desde la base 'default'
        empresa_id = get_current_empresa_id()
        
        empresaDB = EmpresaDB.objects.using('default').get(pk=empresa_id)
                
        if not empresaDB.activa:
            return render(request, 'core/empresa_inactiva.html', {'empresa': empresaDB})
        
        # Leemos la empresa fiscal desde la base del tenant (ya registrada)
        empresa_fiscal = Empresa.objects.using('tenant').first()
        
        if not empresa_fiscal:
            return render(request, 'core/empresa_no_configurada.html', {'empresa': empresaDB})

        contexto = {
            'mensaje1' : config('MENSAJE_INICIO1'),
            'mensaje2' : config('MENSAJE_INICIO2'),
            'mensaje3' : config('MENSAJE_INICIO3'),
        }

        return render(request, 'core/inicio.html', contexto)
    
    except PerfilUsuario.DoesNotExist:
        return redirect('core:login')
    
    except Exception as e:
        # print(f"❌ Error en inicio: {e}")
        return redirect('core:login')

from django.db import connections
def logOutUsuario(request):
    tenant_aliases = list(connections.databases.keys())

    for alias in tenant_aliases:
        if alias != 'default':
            if alias in connections:
                connections[alias].close()
            # Elimina el alias del diccionario de conexiones
            del connections.databases[alias]

    # Limpia la sesión completamente
    request.session.flush() 
    
    # Elimina el tenant actual en thread locals (si usas eso)
    set_current_tenant(None, None, None)

    # Realiza logout
    from django.contrib.auth import logout
    logout(request)
    
    # Redirige donde corresponda
    return redirect('core:login')  # o donde necesites

from django.contrib.auth import authenticate, login
from django.shortcuts import redirect
from core.models import PerfilUsuario

def empresa_inactiva(request):
    return render(request, 'core/empresa_inactiva.html')

def sin_empresa(request):
    return render(request, 'core/sin_empresa.html')

from django.contrib.auth import login
from django.contrib.auth.views import LoginView
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import redirect
from django.urls import reverse
from core.models import PerfilUsuario

class CustomLoginView(LoginView):
    template_name = 'core/login.html'
    form_class = AuthenticationForm

    def form_valid(self, form):
        user = form.get_user()

        try:
            perfil = PerfilUsuario.objects.using('default').get(user=user)
            empresa = EmpresaDB.objects.using('default').get(pk=perfil.empresa_id)
            
            if not empresa or not empresa.activa:
                form.add_error(None, "Empresa inactiva o no asignada.")
                return self.form_invalid(form)

            login(self.request, user)  # 🔒 Esto reinicia la sesión
            update_session_auth_hash(self.request, user)

            # 🔁 Redirige con empresa.id como parámetro
            return redirect(reverse('core:setup_tenant') + f'?eid={empresa.id}')

        except PerfilUsuario.DoesNotExist:
            form.add_error(None, "Usuario o empresa no encontrados.")
            return self.form_invalid(form)

# se llama de Login
@login_required
def setup_tenant(request):
    empresa_id = request.GET.get('eid')
    if not empresa_id:
        # Sin empresa_id, redirigiendo a login.
        return redirect('core:login')

    try:
        # Leer empresa desde la base default
        empresa = EmpresaDB.objects.using('default').get(pk=empresa_id)
        
        # Asegurarse de que hay sesión activa
        if not request.session.session_key:
            request.session.create()

        request.session['empresa_id'] = empresa.id
        request.session['alias_tenant'] = 'tenant'
        request.session['empresa_fiscal'] = None
        request.session.modified = True
        
        # Aquí nos aseguramos que no hay conexiones previas activas
        set_current_tenant('tenant', empresa.id, None)
        
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(reverse('core:inicio'))

    except Exception as e:
        # print(f"❌ Error en setup_tenant: {e}")
        return redirect('core:login')

# FUNCION PARA SIGN INICIAL, AQUI SE CREA LA BD DEL CLIENTE, EL USUARIO INICIAL Y 
# SE AGREGA LA EMPRESA NUEVA EN LA LISTA DE EMPRESAS
from django.shortcuts import render, redirect
from core.services.tenant_setup import crear_tenant_completo
from django.contrib.auth import get_user_model
import traceback

def sign_inicial_view(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre_comercial')
        username = request.POST.get('username')
        password = request.POST.get('password')
        contacto_nombre = request.POST.get('contacto_nombre')
        contacto_telefono = request.POST.get('contacto_telefono')
        contacto_email = request.POST.get('contacto_email')

        if not nombre:
            messages.error(request, f"Se requiere nombre comercial")
            return redirect('core:sign_inicial')

        if not username:
            messages.error(request, f"Se requiere nombre de usuario")
            return redirect('core:sign_inicial')
        
        if not password:
            messages.error(request, f"Se requiere contraseña")
            return redirect('core:sign_inicial')

        if not contacto_nombre:
            messages.error(request, f"Se requiere nombre del contacto")
            return redirect('core:sign_inicial')

        if not contacto_telefono:
            messages.error(request, f"Se requiere teléfono del contacto")
            return redirect('core:sign_inicial')

        if not contacto_email:
            messages.error(request, f"Se requiere email del contacto")
            return redirect('core:sign_inicial')

        UserModel = get_user_model()
        if UserModel.objects.using('default').filter(username=username).exists():
            messages.error(request, f"El nombre de usuario '{username}' ya está en uso.")
            return redirect('core:sign_inicial')
        try:
            crear_tenant_completo(
                request, nombre, username, password,
                contacto_nombre, contacto_telefono, contacto_email
            )
            
            return redirect(f"{reverse('core:login')}?nueva_empresa=1")
        
        except Exception as e:
            print("ERROR EN CREACIÓN DE TENANT:")
            traceback.print_exc()  # Esto imprimirá el error en los logs de docker
            return render(request, 'core/sign_inicial.html', {'error': str(e)})

    return render(request, 'core/sign_inicial.html')

from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render

def es_staff(user):
    return user.is_authenticated and user.is_staff

from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from core.decorators import staff_required

class MenuStaffView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'core/menu_staff.html'

    def test_func(self):
        return self.request.user.is_staff


from django.views.generic import ListView
from django.shortcuts import render
from .models import EmpresaDB
from django.http import HttpResponse
import openpyxl

from django.contrib.admin.views.decorators import staff_member_required

class StaffEmpresaListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = EmpresaDB
    template_name = 'core/staff_empresa_list.html'
    context_object_name = 'empresas'

    def test_func(self):
        return self.request.user.is_staff

    def get_queryset(self):
        return EmpresaDB.objects.using('default').all().order_by('codigo_empresa')

@login_required
@staff_required(redirect_url='core:inicio')
def exportar_empresas_excel(request):
    empresas = EmpresaDB.objects.using('default').all().order_by('codigo_empresa')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empresas"

    # Encabezados
    columnas = ['Código', 'Nombre', 'DB Name', 'Fecha Inicio', 'Fecha Renovación', 'Contacto', 'Teléfono', 'Email']
    ws.append(columnas)

    # Datos
    for emp in empresas:
        ws.append([
            emp.codigo_empresa,
            emp.nombre,
            emp.db_name,
            emp.fecha_inicio.strftime('%Y-%m-%d') if emp.fecha_inicio else '',
            emp.fecha_renovacion.strftime('%Y-%m-%d') if emp.fecha_renovacion else '',
            emp.contacto_nombre,
            emp.contacto_telefono,
            emp.contacto_email,
        ])

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=empresas_registradas.xlsx'
    wb.save(response)
    return response


from timbres.models import TimbresCliente
from django.db.models import F, ExpressionWrapper, BigIntegerField
# Timbres Disponibles por Empresa/Cliente
class StaffTimbresClienteListView22(ListView): 
    model = TimbresCliente
    template_name = 'core/staff_timbres_cliente_list.html'
    context_object_name = 'timbres'

    def get_queryset(self):
        queryset = TimbresCliente.objects.using('default').annotate(
            disponibles_tmp=ExpressionWrapper(
                F('total_asignados') - F('utilizados'),
                output_field=BigIntegerField()
            )
        ).order_by('disponibles_tmp')  # Ordena usando el campo anotado, pero no choca con la propiedad
        return queryset

    def get_context_data(self, **kwargs):
            context = super().get_context_data(**kwargs)
            # Obtener todos los códigos únicos para buscar en EmpresaDB
            codigos = [t.codigo_empresa for t in context['timbres']]
            empresas = EmpresaDB.objects.using('default').filter(codigo_empresa__in=codigos)
            empresas_dict = {e.codigo_empresa: e for e in empresas}
            context['empresas_dict'] = empresas_dict
            return context

from django.views.generic import ListView
from django.db.models import F, ExpressionWrapper, BigIntegerField

class StaffTimbresClienteListView(ListView): 
    model = TimbresCliente
    template_name = 'core/staff_timbres_cliente_list.html'
    context_object_name = 'timbres'

    def get_queryset(self):
        return (
            TimbresCliente.objects.using('default')
            .annotate(
                disponibles_tmp=ExpressionWrapper(
                    F('total_asignados') - F('utilizados'),
                    output_field=BigIntegerField()
                )
            )
            .order_by('disponibles_tmp')
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # self.object_list ya es el queryset devuelto por get_queryset()
        qs = self.object_list

        # Obtener códigos de forma robusta (iterable garantizado)
        codigos = (
            qs.values_list('codigo_empresa', flat=True)
              .distinct()
        )

        # Limpieza adicional por si hay Nones o tipos raros
        codigos_limpios = [str(c) for c in codigos if c not in (None, '')]

        # Usar __in sólo con iterable limpio
        empresas = (
            EmpresaDB.objects.using('default')
            .filter(codigo_empresa__in=codigos_limpios)
        )
        context['empresas_dict'] = {e.codigo_empresa: e for e in empresas}
        return context

@login_required
@staff_required(redirect_url='core:inicio')
def exportar_timbres_disponibles_excel(request):
 
    timbres = (
        TimbresCliente.objects.using('default')
        .annotate(disponibles_temp=ExpressionWrapper(F('total_asignados') - F('utilizados'), output_field=BigIntegerField()))
        .order_by('disponibles_temp')  # Ascendente
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timbres_Disponibles"

    # Encabezados
    columnas = ['Código Empresa','Empresa','Nombre Contacto','Teléfono', 'Email', 'Total Asignados', 'Utilizados', 'Disponibles','Fecha']
    ws.append(columnas)

    # Datos
    # Creamos un diccionario donde cada clave es codigo_empresa y el valor es un diccionario con los demás campos
    empresas = {
        codigo: {
            'nombre': nombre,
            'contacto_nombre': contacto_nombre,
            'contacto_telefono': contacto_telefono,
            'contacto_email': contacto_email
        }
        for codigo, nombre, contacto_nombre, contacto_telefono, contacto_email in 
        EmpresaDB.objects.using('default').values_list(
            'codigo_empresa', 'nombre', 'contacto_nombre', 'contacto_telefono', 'contacto_email'
        )
    }

    # Luego en el for usamos directamente los valores
    for dato in timbres:
        emp = empresas.get(dato.codigo_empresa, {})
        ws.append([
            dato.codigo_empresa,
            emp.get('nombre', ''),
            emp.get('contacto_nombre', ''),
            emp.get('contacto_telefono', ''),
            emp.get('contacto_email', ''),
            dato.total_asignados,
            dato.utilizados,
            dato.total_asignados - dato.utilizados,
            dato.fecha_asignacion.strftime('%Y-%m-%d') if dato.fecha_asignacion else '',
        ])

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=timbres_disponibles.xlsx'
    wb.save(response)
    return response

from django.views.generic import TemplateView
from core.forms import MovimientoTimbresFilterForm
# Timbres asignados por Fecha
class StaffMovimientoTimbresListView(TemplateView):
    template_name = "core/staff_movimiento_timbres_list.html"

    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = MovimientoTimbresFilterForm(self.request.GET or None)
        movimientos = MovimientoTimbresGlobal.objects.using('default').all()
        
        fecha_inicio = None
        fecha_fin = None

        if form.is_valid():
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            movimientos = movimientos.filter(fecha__range=[fecha_inicio, fecha_fin])

            context['fecha_inicio_str'] = fecha_inicio.strftime('%Y-%m-%d') if fecha_inicio else ''
            context['fecha_fin_str'] = fecha_fin.strftime('%Y-%m-%d') if fecha_fin else ''

        else:
            context['fecha_inicio_str'] = localtime(now()).date().strftime('%Y-%m-%d')
            context['fecha_fin_str'] = localtime(now()).date().strftime('%Y-%m-%d')

        # Obtener nombres de empresa
        empresas = EmpresaDB.objects.using('default').in_bulk(field_name='codigo_empresa')
        lista = []
        for m in movimientos:
            m.nombre_empresa = empresas.get(m.codigo_empresa).nombre if empresas.get(m.codigo_empresa) else ''
            lista.append(m)

        context['form'] = form
        context['movimientos'] = lista

        return context

from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.utils.timezone import localdate

# exportar a excel movimientos de timbres en un rango de fechas
@login_required
@staff_required(redirect_url='core:inicio')
def exportar_mov_timbres_excel(request):
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    fecha_inicio = parse_date(fecha_inicio_str) if fecha_inicio_str else None
    fecha_fin = parse_date(fecha_fin_str) if fecha_fin_str else None
    
    hoy = localdate()
    if not fecha_inicio:
        fecha_inicio = hoy
    if not fecha_fin:
        fecha_fin = hoy

    movimientos = MovimientoTimbresGlobal.objects.using('default').filter(
        tipo="S",
        fecha__range=[fecha_inicio, fecha_fin]
        ).order_by('fecha','referencia')
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asignacion de Timbres"

    # Encabezados
    columnas = ['Fecha','Referencia','Empresa', 'Nombre', 'Cantidad', 'Precio Unt', 'Importe', 'Usuario']
    ws.append(columnas)

    # Datos
    empresas = dict(EmpresaDB.objects.using('default').values_list('codigo_empresa', 'nombre'))
    for dato in movimientos:
        empresa_nombre = empresas.get(dato.codigo_empresa, '')
        ws.append([
            dato.fecha.strftime('%Y-%m-%d') if dato.fecha else '',
            dato.referencia,
            dato.codigo_empresa,
            empresa_nombre,
            dato.cantidad,
            dato.precio_unit,
            dato.importe,
            dato.usuario,
        ])

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asignacion_timbres.xlsx'
    wb.save(response)
    return response

class StaffEmpresaRenovacionListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = EmpresaDB
    template_name = 'core/staff_empresa_renovacion_list.html'
    context_object_name = 'empresas'

    def test_func(self):
        return self.request.user.is_staff

    def get_queryset(self):
        return EmpresaDB.objects.using('default').all().order_by('-fecha_renovacion')

@login_required
@staff_required(redirect_url='core:inicio')
def exportar_empresas_renovacion_excel(request):
    empresas = EmpresaDB.objects.using('default').all().order_by('-fecha_renovacion')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empresas_Renovacion"

    # Encabezados
    columnas = ['Código', 'Nombre', 'DB Name', 'Fecha Inicio', 'Fecha Renovación', 'Contacto', 'Teléfono', 'Email']
    ws.append(columnas)

    # Datos
    for emp in empresas:
        ws.append([
            emp.codigo_empresa,
            emp.nombre,
            emp.db_name,
            emp.fecha_inicio.strftime('%Y-%m-%d') if emp.fecha_inicio else '',
            emp.fecha_renovacion.strftime('%Y-%m-%d') if emp.fecha_renovacion else '',
            emp.contacto_nombre,
            emp.contacto_telefono,
            emp.contacto_email,
        ])

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=empresas_renovacion.xlsx'
    wb.save(response)
    return response


from django.views.generic import ListView
from django.http import HttpResponseBadRequest
from django.db.models import Q
from django.contrib.auth.mixins import LoginRequiredMixin

def get_empresa_db_from_tenant(request):
    """
    Dado el tenant actual (alias en sesión), obtiene la Empresa (tenant)
    y luego mapea por codigo_empresa a EmpresaDB (default).
    """
    # Obtén la Empresa en la base tenant. Ajusta el criterio si manejas múltiples.
    empresa_tenant = Empresa.objects.using('tenant').first()

    if not empresa_tenant:
        raise ValueError("No se encontró Empresa activa en la base tenant")

    # Mapea a EmpresaDB en la base default por codigo_empresa
    try:
        empresa_db = EmpresaDB.objects.using('default').get(
            codigo_empresa=empresa_tenant.codigo_empresa
        )
    except EmpresaDB.DoesNotExist:
        raise ValueError(
            f"No existe EmpresaDB con codigo_empresa={empresa_tenant.codigo_empresa}"
        )
    return empresa_db, empresa_tenant


class UsuarioListView(TenantRequiredMixin, ListView):
    """
    Lista usuarios (PerfilUsuario) de la EmpresaDB correspondiente al tenant actual.
    """
    model = PerfilUsuario
    template_name = "core/usuario_list.html"
    context_object_name = "perfiles"
    paginate_by = 25

    def get_queryset(self):
        empresa_db, _ = get_empresa_db_from_tenant(self.request)

        qs = (
            PerfilUsuario.objects.using('default')
            .select_related('user', 'empresa')
            .filter(empresa=empresa_db)
            .order_by('user__username')
        )

        # Búsqueda opcional ?q=
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(
                Q(user__username__icontains=q) |
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q) |
                Q(user__email__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            empresa_db, empresa_tenant = get_empresa_db_from_tenant(self.request)
        except Exception as e:
            # Puedes manejarlo como prefieras
            return {"error": str(e)}

        ctx["empresa_db"] = empresa_db
        ctx["empresa_tenant"] = empresa_tenant
        ctx["q"] = self.request.GET.get("q", "").strip()
        return ctx


from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.db import transaction
from django.shortcuts import render, redirect
from django.utils.crypto import get_random_string

from core.forms import CrearUsuarioEmpresaForm
from core.models import PerfilUsuario
# Si tienes un decorador tenant_required para FBV, puedes usarlo también.

UserModel = get_user_model()

@login_required  
def usuario_create(request):
    # Si tienes decorador tenant_required, colócalo arriba de login_required o debajo (ambos valen)
    # @tenant_required
    try:
        empresa_db, _ = get_empresa_db_from_tenant(request)
    except Exception as e:
        messages.error(request, f"No se pudo determinar la empresa activa: {e}")
        return redirect("core:usuario_list")

    if (empresa_db) and (empresa_db.num_usuarios == 1):
        messages.error(request, "La empresa ya tiene el número máximo de usuarios autorizados.")
        return redirect("core:usuario_list")

    perfil = getattr(request.user, "perfilusuario", None)
    if not perfil or perfil.tipo_usuario != "A":
        messages.error(request, "El Usuario Actual no tiene permisos para crear usuarios.")
        return redirect("core:usuario_list")

    usuarios_activos = (
        PerfilUsuario.objects.using('default')
        .filter(empresa=empresa_db, user__is_active=True)
        .count()
    )
    
    if (empresa_db) and (usuarios_activos >= empresa_db.num_usuarios):
        messages.error(request, "La empresa ya tiene el número máximo de usuarios autorizados.")
        return redirect("core:usuario_list")

    if request.method == "POST":
        form = CrearUsuarioEmpresaForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            email = form.cleaned_data["email"]
            first = form.cleaned_data["first_name"].strip()
            last = form.cleaned_data.get("last_name", "").strip()
            tipo = form.cleaned_data["tipo_usuario"]

            # Re-chequeo por seguridad (la validación primaria ya va en el form)
            if UserModel._default_manager.using('default').filter(username__iexact=username).exists():
                messages.error(request, "El username ya existe, usa otro diferente.")
                return render(request, "core/usuario_form.html", {"form": form})

            temp_password = get_random_string(10)  # más robusto que 7

            try:
                with transaction.atomic(using="default"):
                    user = UserModel._default_manager.db_manager('default').create_user(
                        username=username,
                        email=email,
                        password=temp_password,
                        first_name=first,
                        last_name=last,
                        is_active=True,
                    )
                    PerfilUsuario.objects.using('default').create(
                        user=user,
                        empresa=empresa_db,
                        tipo_usuario=tipo,
                    )
            except Exception as e:
                messages.error(request, f"No se pudo crear el usuario: {e}")
                return render(request, "core/usuario_form.html", {"form": form})

            empresa_nombre = getattr(empresa_db, "nombre_comercial", None) or \
                             getattr(empresa_db, "nombre_fiscal", None) or \
                             getattr(empresa_db, "codigo_empresa", "")

            messages.success(
                request,
                f"Usuario '{username}' creado y ligado a {empresa_nombre}. "
                f"Contraseña temporal: {temp_password}     - favor de guardar estos datos"
            )

            # (Opcional) Forzar cambio de contraseña en primer login:
            # - Puedes guardar un flag en PerfilUsuario (p.ej. must_change_password=True)
            # - O redirigir a una vista para que el admin comparta la clave y el usuario la cambie

            # (Opcional) Enviar correo de invitación:
            # from django.core.mail import EmailMessage
            # from django.conf import settings
            # asunto = "Bienvenido a Switchh"
            # cuerpo = (
            #   f"Hola {first},\n\n"
            #   f"Se te ha dado acceso a {empresa_nombre}.\n"
            #   f"Usuario: {username}\n"
            #   f"Contraseña temporal: {temp_password}\n\n"
            #   "Por favor, inicia sesión y cambia tu contraseña."
            # )
            # EmailMessage(asunto, cuerpo, settings.DEFAULT_FROM_EMAIL, [email]).send()

            return redirect("core:usuario_list")
    else:
        form = CrearUsuarioEmpresaForm()

    return render(request, "core/usuario_form.html", {"form": form})


from django.contrib import messages
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import PasswordChangeView, PasswordChangeDoneView
from django.contrib.auth.forms import SetPasswordForm
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy

from core.forms import BootstrapPasswordChangeForm

UserModel = get_user_model()

class CambiarPasswordView(LoginRequiredMixin, PasswordChangeView):
    template_name = "core/password_change_form.html"
    success_url = reverse_lazy("core:password_change_done")
    # ¡OJO! No definas form_class aquí para que use get_form_class()

    def get_form_class(self):
        target = self._target_user()
        if target.pk == self.request.user.pk:
            # Yo cambio mi propia contraseña -> pide contraseña actual
            return BootstrapPasswordChangeForm
        # Admin cambia a otro -> NO pide contraseña actual
        class BootstrapSetPasswordForm(SetPasswordForm):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                for f in self.fields.values():
                    f.widget.attrs.update({"class": "form-control"})
        return BootstrapSetPasswordForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self._target_user()
        return kwargs

    def form_valid(self, form):
        target = self._target_user()
        # Permisos: admin de la misma empresa (o superuser) para cambiar a otros
        if target.pk != self.request.user.pk and not self._can_admin_change_target(target):
            messages.error(self.request, "No tienes permisos para modificar esta contraseña.")
            return redirect("core:usuario_list")

        resp = super().form_valid(form)

        # Mantener sesión solo si me la cambié a mí mismo
        if target.pk == self.request.user.pk:
            update_session_auth_hash(self.request, target)

        messages.success(self.request, f"Contraseña actualizada para '{target.username}'.")
        return resp

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        target = self._target_user()
        ctx["target_user"] = target
        # Asegura que el POST vaya a la misma URL (con user_id si aplica)
        return ctx

    # ---------- helpers ----------
    def _target_user(self):
        if hasattr(self, "_cached_target_user"):
            return self._cached_target_user

        # 1) Preferir el valor enviado por POST
        uid = self.request.POST.get("target_user_id")
        if uid:
            try:
                uid = int(uid)
            except ValueError:
                uid = None

        # 2) Si no viene en POST, usar el de la URL
        if not uid:
            uid = self.kwargs.get("user_id")

        # 3) Si no hay uid, es el propio usuario
        if not uid:
            self._cached_target_user = self.request.user
        else:
            self._cached_target_user = get_object_or_404(
                UserModel._default_manager.db_manager("default"), pk=uid
            )
        return self._cached_target_user

    def _can_admin_change_target(self, target_user):
        if self.request.user.is_superuser:
            return True
        try:
            empresa_db, _ = get_empresa_db_from_tenant(self.request)
        except Exception:
            return False
        mi_perfil = getattr(self.request.user, "perfilusuario", None)
        if not (mi_perfil and mi_perfil.tipo_usuario == "A" and mi_perfil.empresa_id == empresa_db.id):
            return False
        try:
            perfil_target = PerfilUsuario.objects.using("default").get(user=target_user)
        except PerfilUsuario.DoesNotExist:
            return False
        return perfil_target.empresa_id == empresa_db.id

class CambiarPasswordDoneView(LoginRequiredMixin, PasswordChangeDoneView):
    template_name = "core/password_change_done.html"

# core/views_auth.py (o core/views.py)
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import redirect
from django.http import HttpResponseBadRequest

UserModel = get_user_model()

@login_required
def usuarios_bulk_status(request):
    if request.method != "POST":
        return HttpResponseBadRequest("Método no permitido")

    # 1) Empresa activa + permisos (solo admin de la empresa o superuser)
    try:
        empresa_db, empresa_tenant = get_empresa_db_from_tenant(request)
    except Exception as e:
        messages.error(request, f"No se pudo determinar la empresa activa: {e}")
        return redirect("core:usuario_list")

    mi_perfil = getattr(request.user, "perfilusuario", None)
    if not (request.user.is_superuser or (mi_perfil and mi_perfil.empresa_id == empresa_db.id and mi_perfil.tipo_usuario == "A")):
        messages.error(request, "No tienes permisos para modificar usuarios.")
        return redirect("core:usuario_list")

    # 2) Leer selección y acción
    action = (request.POST.get("action") or "").lower()
    if action not in {"activate", "deactivate"}:
        messages.error(request, "Debes seleccionar una acción válida (activar/desactivar).")
        return redirect("core:usuario_list")

    user_ids = request.POST.getlist("user_ids")
    if not user_ids:
        messages.warning(request, "No seleccionaste usuarios.")
        return redirect("core:usuario_list")

    # Convertir IDs a int válidos
    try:
        user_ids = [int(u) for u in user_ids]
    except ValueError:
        messages.error(request, "Selección inválida.")
        return redirect("core:usuario_list")

    # 3) Limitar a usuarios de ESTA empresa
    perfiles_qs = (
        PerfilUsuario.objects.using("default")
        .select_related("user")
        .filter(empresa=empresa_db, user_id__in=user_ids)
    )
    user_ids_empresa = list(perfiles_qs.values_list("user_id", flat=True))
    if not user_ids_empresa:
        messages.warning(request, "Ningún usuario seleccionado pertenece a esta empresa.")
        return redirect("core:usuario_list")

    # 4) Protecciones básicas
    # - No permitir desactivarte a ti mismo (para no quedarte fuera)
    if action == "deactivate" and request.user.id in user_ids_empresa and not request.user.is_superuser:
        user_ids_empresa.remove(request.user.id)
        messages.info(request, "No puedes desactivar tu propio usuario (omitido).")

    # - No permitir tocar superusers si no eres superuser
    if not request.user.is_superuser:
        superuser_ids = list(
            UserModel._default_manager.db_manager("default")
            .filter(id__in=user_ids_empresa, is_superuser=True)
            .values_list("id", flat=True)
        )
        if superuser_ids:
            user_ids_empresa = [uid for uid in user_ids_empresa if uid not in superuser_ids]
            messages.info(request, "Los superusuarios fueron omitidos.")

    if not user_ids_empresa:
        messages.warning(request, "No hay usuarios válidos para procesar.")
        return redirect("core:usuario_list")

    # 5) Validar cupo al ACTIVAR (Empresa.num_usuarios en tenant)
    if action == "activate":
        limite = empresa_tenant.num_usuarios or 0  # 0/None = sin límite
        if limite:
            usados = (
                PerfilUsuario.objects.using("default")
                .filter(empresa=empresa_db, user__is_active=True)
                .count()
            )
            ya_activos = (
                UserModel._default_manager.db_manager("default")
                .filter(id__in=user_ids_empresa, is_active=True)
                .count()
            )
            a_activar = len(user_ids_empresa) - ya_activos
            if a_activar <= 0:
                messages.info(request, "Todos los usuarios seleccionados ya están activos.")
                return redirect("core:usuario_list")

            if usados + a_activar > limite:
                disponibles = max(limite - usados, 0)
                messages.error(
                    request,
                    f"No se puede activar: límite {usados}/{limite}. "
                    f"Disponibles: {disponibles}. Reduce la selección o aumenta el plan."
                )
                return redirect("core:usuario_list")

    # 6) Ejecutar actualización
    try:
        with transaction.atomic(using="default"):
            qs = UserModel._default_manager.db_manager("default").filter(id__in=user_ids_empresa)
            if action == "activate":
                updated = qs.update(is_active=True)
                messages.success(request, f"Usuarios activados: {updated}.")
            else:
                updated = qs.update(is_active=False)
                messages.success(request, f"Usuarios desactivados: {updated}.")
    except Exception as e:
        messages.error(request, f"No se pudo actualizar el estado de usuarios: {e}")

    return redirect("core:usuario_list")

from django.db.models import Q
from django.urls import reverse
from django.views.generic import ListView, UpdateView
from django.contrib import messages
from django.utils.decorators import method_decorator
from django.contrib.admin.views.decorators import staff_member_required

from core.forms import EmpresaContactoForm

@method_decorator(staff_member_required, name="dispatch")
class EmpresaContactoListView(LoginRequiredMixin, ListView):
    model = EmpresaDB
    template_name = "core/empresa_contacto_list.html"
    context_object_name = "empresas"
    paginate_by = 15

    def get_queryset(self):
        qs = EmpresaDB.objects.using('tenant').all().order_by("codigo_empresa")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(codigo_empresa__icontains=q) |
                Q(nombre__icontains=q) |
                Q(slug__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        return ctx

from django.db import transaction
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse

class EmpresaContactoUpdateView(UpdateView):
    model = EmpresaDB
    form_class = EmpresaContactoForm
    template_name = "core/empresa_contacto_form.html"
    slug_field = "codigo_empresa"
    slug_url_kwarg = "codigo_empresa"

    def get_queryset(self):
        # Leer siempre desde 'default'
        return EmpresaDB.objects.using("default").all()

    def form_valid(self, form):
        obj = self.get_object()  # ya viene de 'default' por get_queryset()

        # Valores limpios desde el form
        fecha = form.cleaned_data.get("fecha_renovacion")
        nombre = form.cleaned_data.get("contacto_nombre")
        tel    = form.cleaned_data.get("contacto_telefono")
        mail   = form.cleaned_data.get("contacto_email")

        with transaction.atomic(using="default"):
            # 🔧 UPDATE directo en la BD 'default'
            updated = (
                EmpresaDB.objects.using("default")
                .filter(pk=obj.pk)
                .update(
                    fecha_renovacion=fecha,
                    contacto_nombre=nombre,
                    contacto_telefono=tel,
                    contacto_email=mail,
                )
            )

        if updated == 0:
            messages.warning(self.request, "No se aplicó ningún cambio (0 filas actualizadas).")
        else:
            messages.success(self.request, "Datos actualizados correctamente.")

        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse("core:empresa_contacto_update", kwargs={
            "codigo_empresa": self.object.codigo_empresa
        })

from core.forms import EmpresaNumUsuariosForm
 
@method_decorator(staff_member_required, name="dispatch")
class EmpresaNumUsuariosListView(LoginRequiredMixin, ListView):
    model = EmpresaDB
    template_name = "core/staff_empresa_num_usuarios_list.html"
    context_object_name = "empresas"
    paginate_by = 15

    def get_queryset(self):
        qs = EmpresaDB.objects.using('default').all().order_by("codigo_empresa")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(codigo_empresa__icontains=q) |
                Q(nombre__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        return ctx

class EmpresaNumUsuariosUpdateView(LoginRequiredMixin, UpdateView):
    model = EmpresaDB
    form_class = EmpresaNumUsuariosForm
    template_name = "core/staff_empresa_num_usuarios_form.html"
    slug_field = "codigo_empresa"
    slug_url_kwarg = "codigo_empresa"

    def get_queryset(self):
        return EmpresaDB.objects.using("default").all()

    def form_valid(self, form):
        obj = self.get_object()  # ya viene de 'default' por get_queryset()

        # Valores limpios desde el form
        numero = form.cleaned_data.get("num_usuarios")

        with transaction.atomic(using="tenant"):
            # 🔧 UPDATE directo en la BD 'default'
            updated = (
                EmpresaDB.objects.using("default")
                .filter(pk=obj.pk)
                .update(
                    num_usuarios=numero,
                )
            )

        if updated == 0:
            messages.warning(self.request, "No se aplicó ningún cambio (0 filas actualizadas).")
        else:
            messages.success(self.request, "Datos actualizados correctamente.")

        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse("core:staff_empresa_num_usuarios_update", kwargs={
            "codigo_empresa": self.object.codigo_empresa
        })
    
